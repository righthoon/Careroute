"""
csv_to_sqlite.py — DUR CSV/XLS → SQLite 변환 스크립트

실행: python csv_to_sqlite.py
결과: dataset/dur.db (배포 시 이 파일만 포함하면 됨)

변환 후 서버는 CSV 대신 SQLite를 사용 → 시작 시간 ~2초
"""

import re
import sqlite3
import time
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).parent))
from config import (
    CSV_ANTIPYRETIC, CSV_CONTRAINDICATED, CSV_ELDERLY, CSV_ENCODING,
    DATASET_DIR, XLS_EFFICACY,
)

DB_PATH = DATASET_DIR / "dur.db"

# ── 정규화 함수 (part2와 동일) ──────────────────────────────────────────

def _vec_normalize(series: pd.Series) -> pd.Series:
    s = series.fillna("").astype(str)
    s = s.str.replace(r"\s+", "", regex=True)
    s = s.str.replace(r"\(.*", "", regex=True)
    s = s.str.replace(r"_.*", "", regex=True)
    s = s.str.replace(
        r"\d[\d/.]*\s*(mg|ml|g|mcg|ug|IU|밀리그램|밀리그람|밀리리터|그람|그램)?",
        "", regex=True, flags=re.IGNORECASE,
    )
    return s.str.lower().str.strip()


def _ingr_first(series: pd.Series) -> pd.Series:
    return series.fillna("").str.lower().str.split().str[0].fillna("")


# ── 변환 ──────────────────────────────────────────────────────────────────

def convert():
    DATASET_DIR.mkdir(exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)

    # 1) 노인주의 (557행)
    t = time.time()
    print("노인주의 CSV 변환 중...")
    df = pd.read_csv(CSV_ELDERLY, encoding=CSV_ENCODING)
    df["_norm"] = _vec_normalize(df["제품명"])
    df["_ingr"] = _ingr_first(df["성분명"])
    df.to_sql("elderly", conn, if_exists="replace", index=False)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_elderly_norm ON elderly(_norm)")
    print(f"  완료: {len(df)}행 ({round(time.time()-t,1)}s)")

    # 2) 노인주의(해열진통소염제) (1034행)
    t = time.time()
    print("해열진통소염제 CSV 변환 중...")
    df = pd.read_csv(CSV_ANTIPYRETIC, encoding=CSV_ENCODING)
    df["_norm"] = _vec_normalize(df["제품명"])
    df["_ingr"] = _ingr_first(df["성분명"])
    df.to_sql("antipyretic", conn, if_exists="replace", index=False)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_antipyretic_norm ON antipyretic(_norm)")
    print(f"  완료: {len(df)}행 ({round(time.time()-t,1)}s)")

    # 3) 병용금기 (873K행) — 핵심 최적화
    t = time.time()
    print("병용금기 CSV 변환 중 (대용량, 1~2분 소요)...")
    df = pd.read_csv(CSV_CONTRAINDICATED, encoding=CSV_ENCODING)
    df["_normA"] = _vec_normalize(df["제품명A"])
    df["_normB"] = _vec_normalize(df["제품명B"])
    df["_ingrA"] = _ingr_first(df["성분명A"])
    df["_ingrB"] = _ingr_first(df["성분명B"])
    df.to_sql("contraindicated", conn, if_exists="replace", index=False,
              chunksize=10000)
    # instr() 검색용 인덱스 (LIKE '%x%'는 인덱스 불가 → 별도 prefix 컬럼 불필요)
    # SQLite instr()는 풀스캔이지만 C레벨이라 pandas보다 10배 빠름
    print(f"  완료: {len(df)}행 ({round(time.time()-t,1)}s)")

    # 4) 효능군 XLS (377행)
    t = time.time()
    print("효능군 XLS 변환 중...")
    with open(XLS_EFFICACY, "rb") as f:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1)]
    xls_df = pd.DataFrame(rows, columns=headers)
    xls_df["_eng"] = _ingr_first(xls_df["DUR성분명영문"])
    xls_df.to_sql("efficacy", conn, if_exists="replace", index=False)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_eff_eng ON efficacy(_eng)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_eff_kor ON efficacy(DUR성분명)")
    print(f"  완료: {len(xls_df)}행 ({round(time.time()-t,1)}s)")

    # 5) 제품→성분 매핑 테이블 (prod_ingr)
    t = time.time()
    print("제품→성분 매핑 테이블 생성 중...")
    prod_ingr = {}
    for tbl in ["elderly", "antipyretic"]:
        cur = conn.execute(f"SELECT _norm, _ingr FROM {tbl}")
        for norm, ingr in cur.fetchall():
            if norm and ingr:
                prod_ingr[norm] = ingr
    # 병용금기 A/B
    cur = conn.execute("SELECT _normA, _ingrA, _normB, _ingrB FROM contraindicated")
    for na, ia, nb, ib in cur.fetchall():
        if na and ia:
            prod_ingr[na] = ia
        if nb and ib:
            prod_ingr[nb] = ib

    pi_df = pd.DataFrame(prod_ingr.items(), columns=["prod_norm", "ingr_first"])
    pi_df.to_sql("prod_ingr", conn, if_exists="replace", index=False)
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_pi_prod ON prod_ingr(prod_norm)")
    print(f"  완료: {len(pi_df)}항목 ({round(time.time()-t,1)}s)")

    conn.commit()
    conn.close()

    size_mb = round(DB_PATH.stat().st_size / 1024 / 1024, 1)
    print(f"\n✅ SQLite 변환 완료: {DB_PATH} ({size_mb}MB)")
    print("서버 시작 시 CSV 대신 이 DB를 사용합니다.")


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    t0 = time.time()
    convert()
    print(f"총 소요시간: {round(time.time()-t0, 1)}s")
