"""
part2_dur_review.py -- DUR Rule-based 약물 리뷰

로딩 우선순위:
  1. dataset/dur.db (SQLite) → 시작 ~2초, 검색 빠름  ← 배포 권장
  2. dataset/dur_db_cache.pkl (pickle 캐시) → 시작 ~2초
  3. 원본 CSV/XLS → 시작 ~30초 (최초 1회 후 pkl 자동 생성)
"""

import pickle
import re
import sqlite3
import warnings
from dataclasses import asdict, dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from config import (
    CSV_ANTIPYRETIC, CSV_CONTRAINDICATED, CSV_ELDERLY, CSV_ENCODING,
    DATASET_DIR, DB_CACHE_PKL, XLS_EFFICACY,
)

warnings.filterwarnings("ignore")

DB_SQLITE = DATASET_DIR / "dur.db"


# ── 데이터 클래스 -----------------------------------------------------------

@dataclass
class Alert:
    alert_type: str
    severity: str
    drugs: list
    detail: str
    matched_product: str
    extra: dict = field(default_factory=dict)

    def to_dict(self):
        return asdict(self)


# ── 정규화 -----------------------------------------------------------------

_RE_SPACE = re.compile(r"\s+")
_RE_PAREN = re.compile(r"\(.*")
_RE_UNDER = re.compile(r"_.*")
_RE_DOSE  = re.compile(
    r"\d[\d/.]*\s*(mg|ml|g|mcg|ug|IU|밀리그램|밀리그람|밀리리터|그람|그램)?",
    re.IGNORECASE,
)


def _normalize(name: str) -> str:
    name = _RE_SPACE.sub("", str(name))
    name = _RE_PAREN.sub("", name)
    name = _RE_UNDER.sub("", name)
    name = _RE_DOSE.sub("", name)
    return name.lower().strip()


def _normalize_ingredient(name: str) -> str:
    if not name:
        return ""
    return str(name).lower().strip().split()[0]


def _vec_normalize(series: pd.Series) -> pd.Series:
    s = series.fillna("").astype(str)
    s = s.str.replace(r"\s+", "", regex=True)
    s = s.str.replace(r"\(.*", "", regex=True)
    s = s.str.replace(r"_.*", "", regex=True)
    s = s.str.replace(
        r"\d[\d/.]*\s*(mg|ml|g|mcg|ug|IU|밀리그램|밀리그람|밀리리터|그람|그램)?",
        "", regex=True, flags=re.IGNORECASE,
    )
    return s.str.lower().str.strip()


def _ingr_first(series: pd.Series) -> pd.Series:
    return series.fillna("").str.lower().str.split().str[0].fillna("")


# ── SQLite 백엔드 -----------------------------------------------------------

class DURDatabaseSQLite:
    """SQLite 기반 DUR DB — 배포 환경 권장"""

    def __init__(self, db_path: Path = DB_SQLITE):
        self.db_path = db_path
        self._conn: Optional[sqlite3.Connection] = None
        self._loaded = False

    def load(self):
        if self._loaded:
            return
        print("DUR DB (SQLite) 로딩 중: " + str(self.db_path))
        self._conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._loaded = True
        cur = self._conn.execute("SELECT COUNT(*) FROM contraindicated")
        count = cur.fetchone()[0]
        print("  SQLite 로드 완료: 병용금기 " + str(count) + "행")

    def _rows_to_dicts(self, rows) -> list:
        return [dict(r) for r in rows]

    def find_elderly(self, query_norm: str) -> list:
        if not query_norm:
            return []
        cur = self._conn.execute(
            "SELECT * FROM elderly WHERE _norm = ? OR instr(_norm, ?) > 0 OR instr(?, _norm) > 0",
            (query_norm, query_norm, query_norm),
        )
        return self._rows_to_dicts(cur.fetchall())

    def find_antipyretic(self, query_norm: str) -> list:
        if not query_norm:
            return []
        cur = self._conn.execute(
            "SELECT * FROM antipyretic WHERE _norm = ? OR instr(_norm, ?) > 0 OR instr(?, _norm) > 0",
            (query_norm, query_norm, query_norm),
        )
        return self._rows_to_dicts(cur.fetchall())

    def find_contraindicated(self, norm_a: str, norm_b: str) -> list:
        if not norm_a or not norm_b:
            return []
        cur = self._conn.execute("""
            SELECT * FROM contraindicated WHERE
              (instr(_normA, ?) > 0 AND instr(_normB, ?) > 0)
           OR (instr(_normA, ?) > 0 AND instr(_normB, ?) > 0)
        """, (norm_a, norm_b, norm_b, norm_a))
        return self._rows_to_dicts(cur.fetchall())

    def find_efficacy_group(self, product_norm: str) -> Optional[str]:
        if not product_norm:
            return None
        # 제품명 → 성분명 첫단어
        cur = self._conn.execute(
            "SELECT ingr_first FROM prod_ingr WHERE prod_norm = ? OR instr(prod_norm, ?) > 0 OR instr(?, prod_norm) > 0 LIMIT 1",
            (product_norm, product_norm, product_norm),
        )
        row = cur.fetchone()
        if not row:
            return None
        ingr = row["ingr_first"]
        # 성분명 → 효능군
        cur2 = self._conn.execute(
            "SELECT 효능군 FROM efficacy WHERE _eng = ? OR DUR성분명 = ? LIMIT 1",
            (ingr, ingr),
        )
        row2 = cur2.fetchone()
        return row2["효능군"] if row2 else None


# ── pickle/CSV 백엔드 (기존 유지) -------------------------------------------

def _cache_is_valid() -> bool:
    if not DB_CACHE_PKL.exists():
        return False
    cache_mtime = DB_CACHE_PKL.stat().st_mtime
    sources = [CSV_ELDERLY, CSV_ANTIPYRETIC, CSV_CONTRAINDICATED, XLS_EFFICACY]
    return all(cache_mtime > src.stat().st_mtime for src in sources if src.exists())


class DURDatabase:
    """CSV/pickle 기반 DUR DB — 로컬 개발 환경"""

    def __init__(self):
        self._loaded = False
        self._elderly_df = None
        self._antipyretic_df = None
        self._contraind_df = None
        self._elderly_index = {}
        self._antipyretic_index = {}
        self._efficacy_index = {}
        self._prod_to_ingr = {}

    def load(self):
        if self._loaded:
            return
        if _cache_is_valid():
            self._load_from_cache()
        else:
            self._load_from_sources()
            self._save_cache()
        self._loaded = True

    def _load_from_cache(self):
        print("DUR DB 캐시 로딩 중...")
        with open(DB_CACHE_PKL, "rb") as f:
            data = pickle.load(f)
        self._elderly_df        = data["elderly_df"]
        self._antipyretic_df    = data["antipyretic_df"]
        self._contraind_df      = data["contraind_df"]
        self._elderly_index     = data["elderly_index"]
        self._antipyretic_index = data["antipyretic_index"]
        self._efficacy_index    = data["efficacy_index"]
        self._prod_to_ingr      = data["prod_to_ingr"]
        print("  캐시 로드 완료: " + str(len(self._contraind_df)) + "행")

    def _load_from_sources(self):
        print("DUR DB 원본 로딩 중 (~30초)...")
        self._elderly_df = pd.read_csv(CSV_ELDERLY, encoding=CSV_ENCODING)
        self._elderly_df["_norm"] = _vec_normalize(self._elderly_df["제품명"])
        self._elderly_index = dict(zip(self._elderly_df["_norm"], self._elderly_df.index))

        self._antipyretic_df = pd.read_csv(CSV_ANTIPYRETIC, encoding=CSV_ENCODING)
        self._antipyretic_df["_norm"] = _vec_normalize(self._antipyretic_df["제품명"])
        self._antipyretic_index = dict(zip(self._antipyretic_df["_norm"], self._antipyretic_df.index))

        print("  병용금기 CSV 로딩 중...")
        self._contraind_df = pd.read_csv(CSV_CONTRAINDICATED, encoding=CSV_ENCODING)
        self._contraind_df["_normA"] = _vec_normalize(self._contraind_df["제품명A"])
        self._contraind_df["_normB"] = _vec_normalize(self._contraind_df["제품명B"])
        self._build_efficacy_and_prod_ingr()
        print("  원본 로드 완료")

    def _save_cache(self):
        DB_CACHE_PKL.parent.mkdir(exist_ok=True)
        with open(DB_CACHE_PKL, "wb") as f:
            pickle.dump({
                "elderly_df": self._elderly_df,
                "antipyretic_df": self._antipyretic_df,
                "contraind_df": self._contraind_df,
                "elderly_index": self._elderly_index,
                "antipyretic_index": self._antipyretic_index,
                "efficacy_index": self._efficacy_index,
                "prod_to_ingr": self._prod_to_ingr,
            }, f)
        print("  캐시 저장: " + str(DB_CACHE_PKL))

    def _build_efficacy_and_prod_ingr(self):
        with open(XLS_EFFICACY, "rb") as f:
            wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
        hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {hdrs[c]: ws.cell(r, c + 1).value for c in range(len(hdrs))}
            eng = _normalize_ingredient(row.get("DUR성분명영문") or "")
            kor = str(row.get("DUR성분명") or "").strip()
            grp = str(row.get("효능군") or "").strip()
            if not grp or grp == "None":
                continue
            if eng:
                self._efficacy_index[eng] = grp
            if kor:
                self._efficacy_index[kor] = grp

        for df in [self._elderly_df, self._antipyretic_df]:
            ingr_s = _ingr_first(df["성분명"])
            self._prod_to_ingr.update(
                {k: v for k, v in zip(df["_norm"], ingr_s) if k and v}
            )
        ingr_a = _ingr_first(self._contraind_df["성분명A"])
        ingr_b = _ingr_first(self._contraind_df["성분명B"])
        self._prod_to_ingr.update(
            {k: v for k, v in zip(self._contraind_df["_normA"], ingr_a) if k and v}
        )
        self._prod_to_ingr.update(
            {k: v for k, v in zip(self._contraind_df["_normB"], ingr_b) if k and v}
        )

    def _match_small(self, query_norm, index, df):
        if query_norm in index:
            return [df.iloc[index[query_norm]]]
        return [df.iloc[idx] for key, idx in index.items()
                if query_norm and (query_norm in key or key in query_norm)]

    def find_elderly(self, query_norm):
        return self._match_small(query_norm, self._elderly_index, self._elderly_df)

    def find_antipyretic(self, query_norm):
        return self._match_small(query_norm, self._antipyretic_index, self._antipyretic_df)

    def find_contraindicated(self, norm_a: str, norm_b: str) -> list:
        if not norm_a or not norm_b:
            return []
        df = self._contraind_df
        hit_a  = df["_normA"].str.contains(norm_a, regex=False, na=False)
        hit_b  = df["_normB"].str.contains(norm_b, regex=False, na=False)
        hit_ba = df["_normA"].str.contains(norm_b, regex=False, na=False)
        hit_ab = df["_normB"].str.contains(norm_a, regex=False, na=False)
        return [row for _, row in df[(hit_a & hit_b) | (hit_ba & hit_ab)].iterrows()]

    def find_efficacy_group(self, product_norm: str) -> Optional[str]:
        ingr = self._prod_to_ingr.get(product_norm)
        if not ingr:
            for key, val in self._prod_to_ingr.items():
                if product_norm and (product_norm in key or key in product_norm):
                    ingr = val
                    break
        return self._efficacy_index.get(ingr) if ingr else None


# ── 자동 백엔드 선택 --------------------------------------------------------

def create_db() -> DURDatabase | DURDatabaseSQLite:
    """SQLite가 있으면 SQLite, 없으면 CSV/pickle 사용"""
    if DB_SQLITE.exists():
        return DURDatabaseSQLite(DB_SQLITE)
    return DURDatabase()


# ── 리뷰어 -----------------------------------------------------------------

class DURReviewer:

    def __init__(self, db=None, patient_age=65):
        self.db = db if db is not None else create_db()
        self.db.load()
        self.age_threshold = patient_age

    @staticmethod
    def _to_envelope_list(drug_lists):
        if not drug_lists:
            return []
        if isinstance(drug_lists[0], str):
            return [drug_lists]
        return drug_lists

    def _collect_drugs(self, envelopes, departments):
        norm_to_orig = {}
        norm_to_dept_keys = {}
        for env_idx, (drugs, dept) in enumerate(zip(envelopes, departments)):
            dept_key = dept.strip() if dept else "__unknown_" + str(env_idx)
            for drug in drugs:
                norm = _normalize(drug)
                if not norm:
                    continue
                if norm not in norm_to_orig:
                    norm_to_orig[norm] = drug
                if norm not in norm_to_dept_keys:
                    norm_to_dept_keys[norm] = set()
                norm_to_dept_keys[norm].add(dept_key)
        return norm_to_orig, norm_to_dept_keys

    def review(self, drug_lists, patient_age=None, departments=None):
        age = patient_age if patient_age is not None else self.age_threshold
        envelopes = self._to_envelope_list(drug_lists)
        depts = list(departments) if departments else [None] * len(envelopes)
        depts += [None] * (len(envelopes) - len(depts))

        norm_to_orig, norm_to_dept_keys = self._collect_drugs(envelopes, depts)
        alerts = []
        seen = set()

        def _add(a):
            key = (a.alert_type, frozenset(a.drugs), a.detail)
            if key not in seen:
                seen.add(key)
                alerts.append(a)

        # Rule 0: 약물중복처방 (다른 과)
        for norm, dept_keys in norm_to_dept_keys.items():
            if len(dept_keys) < 2:
                continue
            orig  = norm_to_orig[norm]
            known = sorted(k for k in dept_keys if not k.startswith("__unknown_"))
            unk   = len(dept_keys) - len(known)
            desc  = ", ".join(known) + (", 과미상 " + str(unk) + "건" if unk else "")
            _add(Alert(
                alert_type="약물중복처방", severity="MEDIUM", drugs=[orig],
                detail="서로 다른 처방에서 동일 약물 중복 (" + desc + ")",
                matched_product=orig,
                extra={"관련과": desc, "처방출처수": str(len(dept_keys))},
            ))

        unique = list(norm_to_orig.items())

        # Rule 1: 병용금기
        n = len(unique)
        for i in range(n):
            for j in range(i + 1, n):
                na, orig_a = unique[i]
                nb, orig_b = unique[j]
                for row in self.db.find_contraindicated(na, nb):
                    _add(Alert(
                        alert_type="병용금기", severity="HIGH",
                        drugs=[orig_a, orig_b],
                        detail=str(row.get("상세정보", "")),
                        matched_product=(str(row.get("제품명A", ""))
                                         + " + " + str(row.get("제품명B", ""))),
                        extra={
                            "성분명A": str(row.get("성분명A", "")),
                            "성분명B": str(row.get("성분명B", "")),
                            "고시번호": str(row.get("고시번호", "")),
                        },
                    ))

        # Rule 2: 효능군중복
        efficacy_map = {}
        for norm, orig in unique:
            grp = self.db.find_efficacy_group(norm)
            if grp:
                efficacy_map.setdefault(grp, []).append(orig)
        for grp, drugs in efficacy_map.items():
            if len(drugs) >= 2:
                _add(Alert(
                    alert_type="효능군중복", severity="MEDIUM", drugs=drugs,
                    detail="동일 효능군(" + grp + ") 약물 " + str(len(drugs)) + "종 중복처방",
                    matched_product=", ".join(drugs),
                    extra={"효능군": grp, "약물수": str(len(drugs))},
                ))

        # Rule 3 & 4: 노인주의
        if age >= self.age_threshold:
            for norm, orig in unique:
                for row in self.db.find_antipyretic(norm):
                    _add(Alert(
                        alert_type="노인주의(해열진통소염제)", severity="MEDIUM",
                        drugs=[orig],
                        detail=str(row.get("약품상세정보", "")),
                        matched_product=str(row.get("제품명", "")),
                        extra={"성분명": str(row.get("성분명", ""))},
                    ))
                for row in self.db.find_elderly(norm):
                    _add(Alert(
                        alert_type="노인주의", severity="MEDIUM",
                        drugs=[orig],
                        detail=str(row.get("약품상세정보", "")),
                        matched_product=str(row.get("제품명", "")),
                        extra={
                            "성분명": str(row.get("성분명", "")),
                            "공고번호": str(row.get("공고번호", "")),
                        },
                    ))

        return alerts

    def format_report(self, alerts):
        if not alerts:
            return "DUR 검토 결과: 이상 없음"
        lines = ["DUR 검토 결과: " + str(len(alerts)) + "건 발견\n"]
        for i, a in enumerate(alerts, 1):
            icon = "[HIGH]" if a.severity == "HIGH" else "[MEDIUM]"
            lines += [
                icon + " [" + str(i) + "] " + a.alert_type,
                "   대상 약물: " + " + ".join(a.drugs),
                "   매칭 제품: " + a.matched_product,
                "   상세 정보: " + a.detail,
            ]
            for k, v in a.extra.items():
                if v and v not in ("nan", "NaN", "None", ""):
                    lines.append("   " + k + ": " + v)
            lines.append("")
        return "\n".join(lines)
